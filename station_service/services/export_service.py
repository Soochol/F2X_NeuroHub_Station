"""
Export Service for Station Service.

Provides export functionality with extensible strategy pattern.
Supports JSON, CSV, Excel (XLSX), and PDF formats.
"""

from __future__ import annotations

import csv
import io
import json
import logging
from abc import ABC, abstractmethod
from datetime import datetime
from typing import Any, Dict, List, Optional, Type, Union

from fastapi.responses import Response
from pydantic import BaseModel

from station_service.api.schemas.report import (
    BatchSummaryReport,
    ExportFormat,
    PeriodStatisticsReport,
    StepAnalysisReport,
)

logger = logging.getLogger(__name__)


def serialize_value(value: Any) -> Any:
    """Serialize a value for JSON/export."""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, BaseModel):
        return value.model_dump()
    return value


class BaseExporter(ABC):
    """Abstract base class for exporters."""

    @property
    @abstractmethod
    def content_type(self) -> str:
        """HTTP Content-Type header value."""
        pass

    @property
    @abstractmethod
    def file_extension(self) -> str:
        """File extension without dot."""
        pass

    @abstractmethod
    def export(self, data: Any, filename_prefix: str) -> Response:
        """
        Export data to the target format.

        Args:
            data: Data to export (Pydantic model or dict).
            filename_prefix: Prefix for the output filename.

        Returns:
            FastAPI Response with the exported data.
        """
        pass

    def _create_response(self, content: bytes | str, filename: str) -> Response:
        """Create a Response with download headers."""
        return Response(
            content=content,
            media_type=self.content_type,
            headers={
                "Content-Disposition": f'attachment; filename="{filename}.{self.file_extension}"'
            },
        )


class JsonExporter(BaseExporter):
    """JSON format exporter."""

    @property
    def content_type(self) -> str:
        return "application/json"

    @property
    def file_extension(self) -> str:
        return "json"

    def export(self, data: Any, filename_prefix: str) -> Response:
        if isinstance(data, BaseModel):
            content = data.model_dump_json(indent=2)
        else:
            content = json.dumps(data, indent=2, default=str)
        return self._create_response(content, filename_prefix)


class CsvExporter(BaseExporter):
    """CSV format exporter."""

    @property
    def content_type(self) -> str:
        return "text/csv; charset=utf-8"

    @property
    def file_extension(self) -> str:
        return "csv"

    def export(self, data: Any, filename_prefix: str) -> Response:
        output = io.StringIO()
        writer = csv.writer(output)

        # Handle different data types
        if isinstance(data, BatchSummaryReport):
            self._export_batch_summary(writer, data)
        elif isinstance(data, PeriodStatisticsReport):
            self._export_period_stats(writer, data)
        elif isinstance(data, StepAnalysisReport):
            self._export_step_analysis(writer, data)
        elif isinstance(data, list):
            self._export_executions(writer, data)
        else:
            # Generic dict export
            self._export_dict(writer, data)

        # Add BOM for Excel compatibility
        content = "\ufeff" + output.getvalue()
        return self._create_response(content.encode("utf-8"), filename_prefix)

    def _export_batch_summary(self, writer: csv.writer, data: BatchSummaryReport) -> None:
        """Export batch summary to CSV."""
        # Summary section
        writer.writerow(["Batch Summary Report"])
        writer.writerow(["Batch ID", data.batch_id])
        writer.writerow(["Batch Name", data.batch_name or ""])
        writer.writerow(["Sequence", f"{data.sequence_name} v{data.sequence_version}"])
        writer.writerow(["Report Generated", data.report_generated_at.isoformat()])
        writer.writerow([])
        writer.writerow(["Overall Statistics"])
        writer.writerow(["Total Executions", data.total_executions])
        writer.writerow(["Pass Count", data.pass_count])
        writer.writerow(["Fail Count", data.fail_count])
        writer.writerow(["Pass Rate (%)", data.pass_rate])
        writer.writerow(["Avg Duration (s)", data.avg_duration])
        writer.writerow([])

        # Steps section
        writer.writerow(["Step Statistics"])
        writer.writerow([
            "Step Name", "Total Runs", "Pass", "Fail",
            "Pass Rate (%)", "Avg Duration (s)", "Min Duration (s)", "Max Duration (s)"
        ])
        for step in data.steps:
            writer.writerow([
                step.step_name, step.total_runs, step.pass_count, step.fail_count,
                step.pass_rate, step.avg_duration, step.min_duration, step.max_duration
            ])

    def _export_period_stats(self, writer: csv.writer, data: PeriodStatisticsReport) -> None:
        """Export period statistics to CSV."""
        writer.writerow(["Period Statistics Report"])
        writer.writerow(["Period Type", data.period_type.value])
        writer.writerow(["Date Range", f"{data.from_date.isoformat()} to {data.to_date.isoformat()}"])
        writer.writerow(["Batch ID", data.batch_id or "All"])
        writer.writerow(["Report Generated", data.report_generated_at.isoformat()])
        writer.writerow([])
        writer.writerow(["Overall Statistics"])
        writer.writerow(["Total Executions", data.total_executions])
        writer.writerow(["Overall Pass Rate (%)", data.overall_pass_rate])
        writer.writerow(["Trend", f"{data.trend_direction} ({data.trend_percentage}%)"])
        writer.writerow([])

        # Period data
        writer.writerow(["Period Data"])
        writer.writerow([
            "Period", "Total", "Pass", "Fail", "Pass Rate (%)", "Avg Duration (s)"
        ])
        for dp in data.data_points:
            writer.writerow([
                dp.period_label, dp.total, dp.pass_count, dp.fail_count,
                dp.pass_rate, dp.avg_duration
            ])

    def _export_step_analysis(self, writer: csv.writer, data: StepAnalysisReport) -> None:
        """Export step analysis to CSV."""
        writer.writerow(["Step Analysis Report"])
        writer.writerow(["Date Range", f"{data.from_date or 'All'} to {data.to_date or 'All'}"])
        writer.writerow(["Batch ID", data.batch_id or "All"])
        writer.writerow(["Report Generated", data.report_generated_at.isoformat()])
        writer.writerow([])
        writer.writerow(["Summary"])
        writer.writerow(["Total Steps Analyzed", data.total_steps])
        writer.writerow(["Most Failed Step", data.most_failed_step or "N/A"])
        writer.writerow(["Slowest Step", data.slowest_step or "N/A"])
        writer.writerow([])

        # Step data
        writer.writerow(["Step Analysis"])
        writer.writerow([
            "Step Name", "Total Runs", "Failures", "Fail Rate (%)",
            "Avg Duration (s)", "P50 Duration (s)", "P95 Duration (s)",
            "Top Failure Reason"
        ])
        for step in data.steps:
            top_reason = step.failure_reasons[0].error_message if step.failure_reasons else ""
            writer.writerow([
                step.step_name, step.total_runs, step.fail_count, step.fail_rate,
                step.avg_duration, step.p50_duration, step.p95_duration,
                top_reason[:100]  # Truncate long error messages
            ])

    def _export_executions(self, writer: csv.writer, data: List[Dict]) -> None:
        """Export execution results to CSV."""
        writer.writerow([
            "Execution ID", "Batch ID", "Sequence", "Version",
            "Status", "Pass", "Duration (s)", "Started At", "Completed At"
        ])
        for exec_data in data:
            writer.writerow([
                exec_data.get("id", ""),
                exec_data.get("batch_id", ""),
                exec_data.get("sequence_name", ""),
                exec_data.get("sequence_version", ""),
                exec_data.get("status", ""),
                exec_data.get("overall_pass", ""),
                exec_data.get("duration", ""),
                exec_data.get("started_at", ""),
                exec_data.get("completed_at", ""),
            ])

    def _export_dict(self, writer: csv.writer, data: Any) -> None:
        """Export generic dict/model to CSV."""
        if isinstance(data, BaseModel):
            data = data.model_dump()
        if isinstance(data, dict):
            for key, value in data.items():
                writer.writerow([key, serialize_value(value)])


class ExcelExporter(BaseExporter):
    """Excel (XLSX) format exporter using openpyxl."""

    @property
    def content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @property
    def file_extension(self) -> str:
        return "xlsx"

    def export(self, data: Any, filename_prefix: str) -> Response:
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise RuntimeError("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=14)

        def style_header_row(worksheet, row_num: int, col_count: int):
            for col in range(1, col_count + 1):
                cell = worksheet.cell(row=row_num, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        def auto_column_width(worksheet):
            for column_cells in worksheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except (TypeError, AttributeError):
                        pass
                worksheet.column_dimensions[column].width = min(max_length + 2, 50)

        # Handle different data types
        if isinstance(data, BatchSummaryReport):
            ws.title = "Summary"
            self._export_batch_summary_xlsx(ws, data, title_font, style_header_row)
        elif isinstance(data, PeriodStatisticsReport):
            ws.title = "Period Stats"
            self._export_period_stats_xlsx(ws, data, title_font, style_header_row)
        elif isinstance(data, StepAnalysisReport):
            ws.title = "Step Analysis"
            self._export_step_analysis_xlsx(ws, data, title_font, style_header_row)
        elif isinstance(data, list):
            ws.title = "Executions"
            self._export_executions_xlsx(ws, data, style_header_row)
        else:
            ws.title = "Data"
            self._export_dict_xlsx(ws, data)

        auto_column_width(ws)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return self._create_response(output.getvalue(), filename_prefix)

    def _export_batch_summary_xlsx(self, ws, data: BatchSummaryReport, title_font, style_header_row):
        """Export batch summary to Excel."""
        row = 1

        # Title
        ws.cell(row=row, column=1, value="Batch Summary Report").font = title_font
        row += 2

        # Summary info
        info_data = [
            ("Batch ID", data.batch_id),
            ("Batch Name", data.batch_name or ""),
            ("Sequence", f"{data.sequence_name} v{data.sequence_version}"),
            ("Report Generated", data.report_generated_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("Total Executions", data.total_executions),
            ("Pass Count", data.pass_count),
            ("Fail Count", data.fail_count),
            ("Pass Rate (%)", data.pass_rate),
            ("Avg Duration (s)", data.avg_duration),
        ]
        for label, value in info_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1

        # Steps table
        ws.cell(row=row, column=1, value="Step Statistics").font = title_font
        row += 1

        headers = ["Step Name", "Total Runs", "Pass", "Fail", "Pass Rate (%)",
                   "Avg Duration (s)", "Min Duration (s)", "Max Duration (s)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        style_header_row(ws, row, len(headers))
        row += 1

        for step in data.steps:
            ws.cell(row=row, column=1, value=step.step_name)
            ws.cell(row=row, column=2, value=step.total_runs)
            ws.cell(row=row, column=3, value=step.pass_count)
            ws.cell(row=row, column=4, value=step.fail_count)
            ws.cell(row=row, column=5, value=step.pass_rate)
            ws.cell(row=row, column=6, value=step.avg_duration)
            ws.cell(row=row, column=7, value=step.min_duration)
            ws.cell(row=row, column=8, value=step.max_duration)
            row += 1

    def _export_period_stats_xlsx(self, ws, data: PeriodStatisticsReport, title_font, style_header_row):
        """Export period statistics to Excel."""
        row = 1

        ws.cell(row=row, column=1, value="Period Statistics Report").font = title_font
        row += 2

        info_data = [
            ("Period Type", data.period_type.value),
            ("Date Range", f"{data.from_date.strftime('%Y-%m-%d')} to {data.to_date.strftime('%Y-%m-%d')}"),
            ("Batch ID", data.batch_id or "All"),
            ("Report Generated", data.report_generated_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("Total Executions", data.total_executions),
            ("Overall Pass Rate (%)", data.overall_pass_rate),
            ("Trend", f"{data.trend_direction} ({data.trend_percentage}%)"),
        ]
        for label, value in info_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1

        # Period data table
        headers = ["Period", "Total", "Pass", "Fail", "Pass Rate (%)", "Avg Duration (s)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        style_header_row(ws, row, len(headers))
        row += 1

        for dp in data.data_points:
            ws.cell(row=row, column=1, value=dp.period_label)
            ws.cell(row=row, column=2, value=dp.total)
            ws.cell(row=row, column=3, value=dp.pass_count)
            ws.cell(row=row, column=4, value=dp.fail_count)
            ws.cell(row=row, column=5, value=dp.pass_rate)
            ws.cell(row=row, column=6, value=dp.avg_duration)
            row += 1

    def _export_step_analysis_xlsx(self, ws, data: StepAnalysisReport, title_font, style_header_row):
        """Export step analysis to Excel."""
        row = 1

        ws.cell(row=row, column=1, value="Step Analysis Report").font = title_font
        row += 2

        info_data = [
            ("Date Range", f"{data.from_date or 'All'} to {data.to_date or 'All'}"),
            ("Batch ID", data.batch_id or "All"),
            ("Report Generated", data.report_generated_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("Total Steps Analyzed", data.total_steps),
            ("Most Failed Step", data.most_failed_step or "N/A"),
            ("Slowest Step", data.slowest_step or "N/A"),
        ]
        for label, value in info_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=str(value) if value else "")
            row += 1

        row += 1

        # Step analysis table
        headers = ["Step Name", "Total Runs", "Failures", "Fail Rate (%)",
                   "Avg Duration (s)", "P50 (s)", "P95 (s)", "Top Failure Reason"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        style_header_row(ws, row, len(headers))
        row += 1

        for step in data.steps:
            top_reason = step.failure_reasons[0].error_message[:100] if step.failure_reasons else ""
            ws.cell(row=row, column=1, value=step.step_name)
            ws.cell(row=row, column=2, value=step.total_runs)
            ws.cell(row=row, column=3, value=step.fail_count)
            ws.cell(row=row, column=4, value=step.fail_rate)
            ws.cell(row=row, column=5, value=step.avg_duration)
            ws.cell(row=row, column=6, value=step.p50_duration)
            ws.cell(row=row, column=7, value=step.p95_duration)
            ws.cell(row=row, column=8, value=top_reason)
            row += 1

    def _export_executions_xlsx(self, ws, data: List[Dict], style_header_row):
        """Export execution results to Excel."""
        headers = ["Execution ID", "Batch ID", "Sequence", "Version",
                   "Status", "Pass", "Duration (s)", "Started At", "Completed At"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        style_header_row(ws, 1, len(headers))

        for row_num, exec_data in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=exec_data.get("id", ""))
            ws.cell(row=row_num, column=2, value=exec_data.get("batch_id", ""))
            ws.cell(row=row_num, column=3, value=exec_data.get("sequence_name", ""))
            ws.cell(row=row_num, column=4, value=exec_data.get("sequence_version", ""))
            ws.cell(row=row_num, column=5, value=exec_data.get("status", ""))
            ws.cell(row=row_num, column=6, value="Pass" if exec_data.get("overall_pass") else "Fail")
            ws.cell(row=row_num, column=7, value=exec_data.get("duration", 0))
            ws.cell(row=row_num, column=8, value=exec_data.get("started_at", ""))
            ws.cell(row=row_num, column=9, value=exec_data.get("completed_at", ""))

    def _export_dict_xlsx(self, ws, data: Any):
        """Export generic dict to Excel."""
        if isinstance(data, BaseModel):
            data = data.model_dump()
        if isinstance(data, dict):
            row = 1
            for key, value in data.items():
                ws.cell(row=row, column=1, value=str(key))
                ws.cell(row=row, column=2, value=str(serialize_value(value)))
                row += 1


class PdfExporter(BaseExporter):
    """PDF format exporter using reportlab."""

    @property
    def content_type(self) -> str:
        return "application/pdf"

    @property
    def file_extension(self) -> str:
        return "pdf"

    def export(self, data: Any, filename_prefix: str) -> Response:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError:
            logger.error("reportlab not installed. Install with: pip install reportlab")
            raise RuntimeError("reportlab not installed")

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=15 * mm,
            leftMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=16,
            spaceAfter=12,
        )
        subtitle_style = ParagraphStyle(
            "Subtitle",
            parent=styles["Heading2"],
            fontSize=12,
            spaceAfter=8,
        )

        elements = []

        # Handle different data types
        if isinstance(data, BatchSummaryReport):
            self._build_batch_summary_pdf(elements, data, title_style, subtitle_style, colors)
        elif isinstance(data, PeriodStatisticsReport):
            self._build_period_stats_pdf(elements, data, title_style, subtitle_style, colors)
        elif isinstance(data, StepAnalysisReport):
            self._build_step_analysis_pdf(elements, data, title_style, subtitle_style, colors)
        elif isinstance(data, list):
            self._build_executions_pdf(elements, data, title_style, colors)
        else:
            elements.append(Paragraph("Report Data", title_style))
            elements.append(Paragraph(str(data), styles["Normal"]))

        doc.build(elements)
        output.seek(0)

        return self._create_response(output.getvalue(), filename_prefix)

    def _create_table_style(self, colors):
        """Create standard table style."""
        return TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])

    def _build_batch_summary_pdf(self, elements, data: BatchSummaryReport, title_style, subtitle_style, colors):
        """Build batch summary PDF content."""
        from reportlab.platypus import Paragraph, Spacer, Table

        elements.append(Paragraph("Batch Summary Report", title_style))
        elements.append(Spacer(1, 12))

        # Summary info
        info_data = [
            ["Batch ID", data.batch_id, "Sequence", f"{data.sequence_name} v{data.sequence_version}"],
            ["Total Executions", str(data.total_executions), "Pass Rate", f"{data.pass_rate}%"],
            ["Pass Count", str(data.pass_count), "Fail Count", str(data.fail_count)],
            ["Avg Duration", f"{data.avg_duration}s", "Report Date", data.report_generated_at.strftime("%Y-%m-%d %H:%M")],
        ]
        info_table = Table(info_data, colWidths=[80, 120, 80, 120])
        info_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8E8E8")),
            ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#E8E8E8")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 20))

        # Steps table
        if data.steps:
            elements.append(Paragraph("Step Statistics", subtitle_style))
            step_data = [["Step Name", "Total", "Pass", "Fail", "Pass Rate", "Avg Duration"]]
            for step in data.steps:
                step_data.append([
                    step.step_name[:30],
                    str(step.total_runs),
                    str(step.pass_count),
                    str(step.fail_count),
                    f"{step.pass_rate}%",
                    f"{step.avg_duration}s",
                ])
            step_table = Table(step_data, colWidths=[150, 60, 60, 60, 70, 80])
            step_table.setStyle(self._create_table_style(colors))
            elements.append(step_table)

    def _build_period_stats_pdf(self, elements, data: PeriodStatisticsReport, title_style, subtitle_style, colors):
        """Build period statistics PDF content."""
        from reportlab.platypus import Paragraph, Spacer, Table

        elements.append(Paragraph("Period Statistics Report", title_style))
        elements.append(Spacer(1, 12))

        info_data = [
            ["Period Type", data.period_type.value, "Date Range",
             f"{data.from_date.strftime('%Y-%m-%d')} to {data.to_date.strftime('%Y-%m-%d')}"],
            ["Total Executions", str(data.total_executions), "Pass Rate", f"{data.overall_pass_rate}%"],
            ["Trend", f"{data.trend_direction} ({data.trend_percentage}%)", "Batch", data.batch_id or "All"],
        ]
        info_table = Table(info_data, colWidths=[80, 100, 80, 140])
        info_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8E8E8")),
            ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#E8E8E8")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 20))

        # Period data table
        if data.data_points:
            elements.append(Paragraph("Period Data", subtitle_style))
            period_data = [["Period", "Total", "Pass", "Fail", "Pass Rate", "Avg Duration"]]
            for dp in data.data_points:
                period_data.append([
                    dp.period_label,
                    str(dp.total),
                    str(dp.pass_count),
                    str(dp.fail_count),
                    f"{dp.pass_rate}%",
                    f"{dp.avg_duration}s",
                ])
            period_table = Table(period_data, colWidths=[100, 60, 60, 60, 70, 80])
            period_table.setStyle(self._create_table_style(colors))
            elements.append(period_table)

    def _build_step_analysis_pdf(self, elements, data: StepAnalysisReport, title_style, subtitle_style, colors):
        """Build step analysis PDF content."""
        from reportlab.platypus import Paragraph, Spacer, Table

        elements.append(Paragraph("Step Analysis Report", title_style))
        elements.append(Spacer(1, 12))

        info_data = [
            ["Total Steps", str(data.total_steps), "Most Failed", data.most_failed_step or "N/A"],
            ["Slowest Step", data.slowest_step or "N/A", "Batch", data.batch_id or "All"],
        ]
        info_table = Table(info_data, colWidths=[80, 140, 80, 140])
        info_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8E8E8")),
            ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#E8E8E8")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 20))

        # Step analysis table
        if data.steps:
            elements.append(Paragraph("Step Details", subtitle_style))
            step_data = [["Step", "Runs", "Fails", "Fail Rate", "Avg", "P50", "P95"]]
            for step in data.steps:
                step_data.append([
                    step.step_name[:25],
                    str(step.total_runs),
                    str(step.fail_count),
                    f"{step.fail_rate}%",
                    f"{step.avg_duration}s",
                    f"{step.p50_duration}s",
                    f"{step.p95_duration}s",
                ])
            step_table = Table(step_data, colWidths=[140, 50, 50, 60, 60, 60, 60])
            step_table.setStyle(self._create_table_style(colors))
            elements.append(step_table)

    def _build_executions_pdf(self, elements, data: List[Dict], title_style, colors):
        """Build executions PDF content."""
        from reportlab.platypus import Paragraph, Spacer, Table

        elements.append(Paragraph("Execution Results", title_style))
        elements.append(Spacer(1, 12))

        exec_data = [["ID", "Batch", "Sequence", "Status", "Pass", "Duration"]]
        for exec_item in data[:50]:  # Limit rows for PDF
            exec_data.append([
                exec_item.get("id", "")[:20],
                exec_item.get("batch_id", "")[:15],
                exec_item.get("sequence_name", "")[:20],
                exec_item.get("status", ""),
                "Pass" if exec_item.get("overall_pass") else "Fail",
                f"{exec_item.get('duration', 0)}s",
            ])
        exec_table = Table(exec_data, colWidths=[100, 80, 100, 70, 50, 60])
        exec_table.setStyle(self._create_table_style(colors))
        elements.append(exec_table)


class ExportService:
    """
    Export Service with strategy pattern for extensibility.

    Usage:
        service = ExportService()
        exporter = service.get_exporter(ExportFormat.XLSX)
        response = exporter.export(data, "report_2025-01-01")

    Extending with new formats:
        class MyExporter(BaseExporter):
            ...

        ExportService.register_exporter(ExportFormat.MY_FORMAT, MyExporter)
    """

    _exporters: Dict[ExportFormat, Type[BaseExporter]] = {
        ExportFormat.JSON: JsonExporter,
        ExportFormat.CSV: CsvExporter,
        ExportFormat.XLSX: ExcelExporter,
        ExportFormat.PDF: PdfExporter,
    }

    def get_exporter(self, format: ExportFormat) -> BaseExporter:
        """
        Get exporter for the specified format.

        Args:
            format: Export format.

        Returns:
            Exporter instance.

        Raises:
            ValueError: If format is not registered.
        """
        exporter_class = self._exporters.get(format)
        if not exporter_class:
            raise ValueError(f"Unknown export format: {format}")
        return exporter_class()

    @classmethod
    def register_exporter(
        cls,
        format: ExportFormat,
        exporter_class: Type[BaseExporter],
    ) -> None:
        """
        Register a new exporter.

        Args:
            format: Export format enum value.
            exporter_class: Exporter class to register.
        """
        cls._exporters[format] = exporter_class
        logger.info(f"Registered exporter: {format} -> {exporter_class.__name__}")

    @classmethod
    def available_formats(cls) -> List[ExportFormat]:
        """Get list of available export formats."""
        return list(cls._exporters.keys())
