"""
Unit tests for ExportService and exporters.
"""

import io
import json
import pytest
from datetime import datetime
from unittest.mock import MagicMock, patch

from fastapi.responses import Response

from station_service.api.schemas.report import (
    ExportFormat,
    BatchSummaryReport,
    StepSummary,
    PeriodStatisticsReport,
    PeriodDataPoint,
    StepAnalysisReport,
    StepAnalysisItem,
    FailureReason,
)
from station_service.services.export_service import (
    ExportService,
    JsonExporter,
    CsvExporter,
    ExcelExporter,
    PdfExporter,
)


# ============================================================================
# Fixtures
# ============================================================================


@pytest.fixture
def sample_batch_summary_report():
    """Create a sample batch summary report."""
    return BatchSummaryReport(
        batch_id="batch-001",
        batch_name="Test Batch",
        sequence_name="test_sequence",
        sequence_version="1.0.0",
        report_generated_at=datetime.now(),
        total_executions=100,
        pass_count=85,
        fail_count=15,
        pass_rate=0.85,
        avg_duration=5500.0,
        steps=[
            StepSummary(
                step_name="Initialize",
                total_runs=100,
                pass_count=98,
                fail_count=2,
                pass_rate=0.98,
                avg_duration=1200.0,
                min_duration=800.0,
                max_duration=2500.0,
            ),
            StepSummary(
                step_name="Execute",
                total_runs=100,
                pass_count=90,
                fail_count=10,
                pass_rate=0.90,
                avg_duration=3500.0,
                min_duration=1500.0,
                max_duration=8000.0,
            ),
        ],
        first_execution=datetime.now(),
        last_execution=datetime.now(),
    )


@pytest.fixture
def sample_period_stats_report():
    """Create a sample period statistics report."""
    return PeriodStatisticsReport(
        period_type="daily",
        from_date=datetime.now(),
        to_date=datetime.now(),
        batch_id=None,
        report_generated_at=datetime.now(),
        total_executions=100,
        overall_pass_rate=0.85,
        data_points=[
            PeriodDataPoint(
                period_start=datetime.now(),
                period_end=datetime.now(),
                period_label="2025-01-01",
                total=20,
                pass_count=18,
                fail_count=2,
                pass_rate=0.90,
                avg_duration=5000.0,
            ),
            PeriodDataPoint(
                period_start=datetime.now(),
                period_end=datetime.now(),
                period_label="2025-01-02",
                total=25,
                pass_count=20,
                fail_count=5,
                pass_rate=0.80,
                avg_duration=5500.0,
            ),
        ],
        trend_direction="stable",
        trend_percentage=0.0,
    )


@pytest.fixture
def sample_step_analysis_report():
    """Create a sample step analysis report."""
    return StepAnalysisReport(
        from_date=None,
        to_date=None,
        batch_id=None,
        report_generated_at=datetime.now(),
        steps=[
            StepAnalysisItem(
                step_name="Execute Test",
                total_runs=100,
                fail_count=10,
                fail_rate=0.10,
                avg_duration=3500.0,
                min_duration=1500.0,
                max_duration=8000.0,
                p50_duration=3200.0,
                p95_duration=7000.0,
                failure_reasons=[
                    FailureReason(
                        error_message="Timeout exceeded",
                        occurrence_count=5,
                        percentage=0.50,
                    ),
                    FailureReason(
                        error_message="Connection failed",
                        occurrence_count=3,
                        percentage=0.30,
                    ),
                ],
            ),
        ],
        total_steps=1,
        most_failed_step="Execute Test",
        slowest_step="Execute Test",
    )


# ============================================================================
# ExportService Tests
# ============================================================================


class TestExportService:
    """Tests for ExportService factory."""

    def test_get_exporter_json(self):
        """Test getting JsonExporter."""
        service = ExportService()
        exporter = service.get_exporter(ExportFormat.JSON)

        assert isinstance(exporter, JsonExporter)

    def test_get_exporter_csv(self):
        """Test getting CsvExporter."""
        service = ExportService()
        exporter = service.get_exporter(ExportFormat.CSV)

        assert isinstance(exporter, CsvExporter)

    def test_get_exporter_xlsx(self):
        """Test getting ExcelExporter."""
        service = ExportService()
        exporter = service.get_exporter(ExportFormat.XLSX)

        assert isinstance(exporter, ExcelExporter)

    def test_get_exporter_pdf(self):
        """Test getting PdfExporter."""
        service = ExportService()
        exporter = service.get_exporter(ExportFormat.PDF)

        assert isinstance(exporter, PdfExporter)

    def test_get_exporter_invalid_format(self):
        """Test getting exporter with invalid format raises ValueError."""
        service = ExportService()

        with pytest.raises(ValueError, match="Unknown export format"):
            service.get_exporter("invalid_format")

    def test_register_custom_exporter(self):
        """Test registering a custom exporter."""
        class CustomExporter:
            def export(self, data, filename_prefix):
                return Response(content="custom", media_type="text/plain")

        ExportService.register_exporter("custom", CustomExporter)

        service = ExportService()
        exporter = service.get_exporter("custom")

        assert isinstance(exporter, CustomExporter)

        # Cleanup
        del ExportService._exporters["custom"]


# ============================================================================
# JsonExporter Tests
# ============================================================================


class TestJsonExporter:
    """Tests for JsonExporter."""

    def test_export_batch_summary(self, sample_batch_summary_report):
        """Test exporting batch summary to JSON."""
        exporter = JsonExporter()
        response = exporter.export(sample_batch_summary_report, "batch_summary")

        assert response.media_type == "application/json"
        assert "batch_summary" in response.headers.get("content-disposition", "")

        # Verify JSON is valid
        content = json.loads(response.body)
        assert content["batch_id"] == "batch-001"
        assert content["total_executions"] == 100

    def test_export_period_stats(self, sample_period_stats_report):
        """Test exporting period stats to JSON."""
        exporter = JsonExporter()
        response = exporter.export(sample_period_stats_report, "period_stats")

        assert response.media_type == "application/json"

        content = json.loads(response.body)
        assert content["total_executions"] == 100
        assert len(content["data_points"]) == 2

    def test_export_step_analysis(self, sample_step_analysis_report):
        """Test exporting step analysis to JSON."""
        exporter = JsonExporter()
        response = exporter.export(sample_step_analysis_report, "step_analysis")

        assert response.media_type == "application/json"

        content = json.loads(response.body)
        assert content["total_steps"] == 1
        assert content["most_failed_step"] == "Execute Test"


# ============================================================================
# CsvExporter Tests
# ============================================================================


class TestCsvExporter:
    """Tests for CsvExporter."""

    def test_export_batch_summary(self, sample_batch_summary_report):
        """Test exporting batch summary to CSV."""
        exporter = CsvExporter()
        response = exporter.export(sample_batch_summary_report, "batch_summary")

        assert response.media_type == "text/csv"
        assert ".csv" in response.headers.get("content-disposition", "")

        # Verify CSV content
        content = response.body.decode("utf-8")
        assert "step_name" in content.lower() or "Step Name" in content
        assert "Initialize" in content
        assert "Execute" in content

    def test_export_period_stats(self, sample_period_stats_report):
        """Test exporting period stats to CSV."""
        exporter = CsvExporter()
        response = exporter.export(sample_period_stats_report, "period_stats")

        assert response.media_type == "text/csv"

        content = response.body.decode("utf-8")
        assert "2025-01-01" in content
        assert "2025-01-02" in content

    def test_export_step_analysis(self, sample_step_analysis_report):
        """Test exporting step analysis to CSV."""
        exporter = CsvExporter()
        response = exporter.export(sample_step_analysis_report, "step_analysis")

        assert response.media_type == "text/csv"

        content = response.body.decode("utf-8")
        assert "Execute Test" in content


# ============================================================================
# ExcelExporter Tests
# ============================================================================


class TestExcelExporter:
    """Tests for ExcelExporter."""

    def test_export_batch_summary(self, sample_batch_summary_report):
        """Test exporting batch summary to Excel."""
        exporter = ExcelExporter()
        response = exporter.export(sample_batch_summary_report, "batch_summary")

        assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert ".xlsx" in response.headers.get("content-disposition", "")

        # Verify it's a valid Excel file by checking magic bytes
        assert response.body[:4] == b"PK\x03\x04"  # ZIP magic number (xlsx is a zip)

    def test_export_period_stats(self, sample_period_stats_report):
        """Test exporting period stats to Excel."""
        exporter = ExcelExporter()
        response = exporter.export(sample_period_stats_report, "period_stats")

        assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert response.body[:4] == b"PK\x03\x04"

    def test_export_step_analysis(self, sample_step_analysis_report):
        """Test exporting step analysis to Excel."""
        exporter = ExcelExporter()
        response = exporter.export(sample_step_analysis_report, "step_analysis")

        assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert response.body[:4] == b"PK\x03\x04"

    def test_export_creates_valid_workbook(self, sample_batch_summary_report):
        """Test that exported Excel is a valid workbook."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter()
        response = exporter.export(sample_batch_summary_report, "test")

        # Load the workbook from bytes
        wb = load_workbook(io.BytesIO(response.body))
        assert wb.active is not None
        assert len(wb.sheetnames) >= 1


# ============================================================================
# PdfExporter Tests
# ============================================================================


class TestPdfExporter:
    """Tests for PdfExporter."""

    def test_export_batch_summary(self, sample_batch_summary_report):
        """Test exporting batch summary to PDF."""
        exporter = PdfExporter()
        response = exporter.export(sample_batch_summary_report, "batch_summary")

        assert response.media_type == "application/pdf"
        assert ".pdf" in response.headers.get("content-disposition", "")

        # Verify it's a valid PDF by checking magic bytes
        assert response.body[:4] == b"%PDF"

    def test_export_period_stats(self, sample_period_stats_report):
        """Test exporting period stats to PDF."""
        exporter = PdfExporter()
        response = exporter.export(sample_period_stats_report, "period_stats")

        assert response.media_type == "application/pdf"
        assert response.body[:4] == b"%PDF"

    def test_export_step_analysis(self, sample_step_analysis_report):
        """Test exporting step analysis to PDF."""
        exporter = PdfExporter()
        response = exporter.export(sample_step_analysis_report, "step_analysis")

        assert response.media_type == "application/pdf"
        assert response.body[:4] == b"%PDF"

    def test_export_creates_readable_pdf(self, sample_batch_summary_report):
        """Test that exported PDF is readable."""
        exporter = PdfExporter()
        response = exporter.export(sample_batch_summary_report, "test")

        # Basic PDF structure validation
        content = response.body
        assert b"%PDF" in content
        assert b"%%EOF" in content or b"endobj" in content


# ============================================================================
# Integration Tests
# ============================================================================


class TestExportServiceIntegration:
    """Integration tests for export workflow."""

    def test_full_export_workflow_json(self, sample_batch_summary_report):
        """Test complete JSON export workflow."""
        service = ExportService()
        exporter = service.get_exporter(ExportFormat.JSON)
        response = exporter.export(sample_batch_summary_report, "test")

        # Verify response is valid
        assert response.status_code == 200
        data = json.loads(response.body)
        assert data["batch_id"] == "batch-001"

    def test_full_export_workflow_xlsx(self, sample_batch_summary_report):
        """Test complete Excel export workflow."""
        service = ExportService()
        exporter = service.get_exporter(ExportFormat.XLSX)
        response = exporter.export(sample_batch_summary_report, "test")

        # Verify response is valid
        assert response.status_code == 200
        assert len(response.body) > 0

    def test_full_export_workflow_pdf(self, sample_batch_summary_report):
        """Test complete PDF export workflow."""
        service = ExportService()
        exporter = service.get_exporter(ExportFormat.PDF)
        response = exporter.export(sample_batch_summary_report, "test")

        # Verify response is valid
        assert response.status_code == 200
        assert len(response.body) > 0
